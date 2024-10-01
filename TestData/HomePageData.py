import openpyxl


class HomePageData:

    #test_HomePage_data = [{"firstName" : "Srinivasa", "email" : "srinivasa@gmail.com", "gender" : "Male"}, {"firstName" : "Mrunal", "email" : "mrunal@gmail.com", "gender" : "Female"}]
    @staticmethod
    def getTestData(testcase_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\kondsrin\\PycharmProjects\\pytestframework_protocommerce\\TestData\\excelDemoPython.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
