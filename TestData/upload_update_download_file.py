import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

file_path = "C:\\Users\\kondsrin\\Downloads\\download (1).xlsx"
dict = {}
fruit_name = 'Apple'
new_price_for_fruit = '500'
column_name = 'price'
toast_loactor = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")

def update_excel_data(file_path, target_fruit, col_name, new_fruitPrice):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    # Identify the required fruit cell
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == target_fruit:
                dict["row"] = i
                break
    # Identify the required column cell
    for j in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=j).value == col_name:
            dict["col"] = j
            break
    sheet.cell(row=dict["row"], column=dict["col"]).value = new_fruitPrice
    # sheet["D3"] = new_price
    book.save(file_path)

driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

# Doanload the file
driver.find_element(By.ID, "downloadButton").click()
time.sleep(5)

# Upadte the table data
update_excel_data(file_path, fruit_name, column_name, new_price_for_fruit)

# upload a file to webPage
driver.find_element(By.CSS_SELECTOR, "input[type='file']").send_keys(file_path)
wait = WebDriverWait(driver, 10)
wait.until(EC.visibility_of_element_located(toast_loactor))
print(driver.find_element(*toast_loactor).text)
wait.until(EC.invisibility_of_element_located(toast_loactor))

price_column_number = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
#By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[{0}]/div".format(str(price_column_number))
price_locator = (By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div["+price_column_number+"]/div")
updated_price = driver.find_element(*price_locator).text

assert (updated_price == new_price_for_fruit)



