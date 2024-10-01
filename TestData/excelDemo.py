import openpyxl

book = openpyxl.load_workbook("C:\\Users\\kondsrin\\PycharmProjects\\pytestframework_protocommerce\\TestData\\excelDemoPython.xlsx")
sheet = book.active

# cell = sheet.cell(row=2, column=2)
# print(cell.value)
#
# sheet.cell(row=2, column=2).value = "Raju"
#
# print(sheet.cell(row=2, column=2).value)

# different method

Dict = {}
cell = sheet["B3"]
print(cell.value)

sheet["A6"].value = "T6"

print(sheet["A6"].value)
print(sheet.max_row)
print(sheet.max_column)
# k = 1
# for char in range(65, 91):
#     col = f'{chr(char)}'
#     if k == sheet.max_column+1:
#         break
#     else:
#         for i in range(1, sheet.max_row + 1):
#             print(sheet["{}{}".format(col,i)].value)
#     k = k + 1


for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "test2":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
